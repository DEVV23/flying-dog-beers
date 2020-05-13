import base64
import io
from app import app
from app import server
import dash_html_components as html
import dash_core_components as dcc
import dash_table_experiments as dte
from dash.dependencies import Input, Output, State
from datetime import datetime
import pandas as pd
import numpy as np
import dash
import dash_table
import os
from openpyxl.utils.dataframe import dataframe_to_rows

from PriceIndices import MarketHistory, Indices
from openpyxl import load_workbook
from os import listdir
from os.path import isfile, join

#history = MarketHistory()
#df = history.get_price('bitcoin', '20130428', '20200510')  # Get Bitcoin price data
#df['date'] = pd.to_datetime(df['date'])
#colors = {
#    'background': '#111111',
#    'background2': '#FF0',
#    'text': 'yellow'
#    }
pathway = 'C:\\Users\\48022\\Documents\\Glassdor\\Survey Automation\\Files'

files = [f for f in listdir(pathway) if isfile(join(pathway, f))]
#firefox_path = 'C:/Program Files/Mozilla Firefox/firefox.exe %s'


#webbrowser.get("http://127.0.0.1:8050/").open(url)
#url = 'https://www.youtube.com/'
#webbrowser.get(firefox_path)
#print (webbrowser._browsers)

#book = load_workbook('Dynata_v4.xlsx')
#writer = pd.ExcelWriter('Dynata_v4.xlsx', engine='openpyxl') 
#writer.book = book
#
### ExcelWriter for some reason uses writer.sheets to access the sheet.
### If you leave it empty it will not know that sheet Main is already there
### and will create a new sheet.
#
#writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
#final_datamap_df.to_excel(writer, "UpdatedDataMap",index = False,header = False)
#writer.save()



app = dash.Dash(__name__)

external_stylesheets = ['https://codepen.io/chriddyp/pen/bWLwgP.css']
app = dash.Dash(__name__, external_stylesheets=external_stylesheets)
server = app.server
app.config.suppress_callback_exceptions = True


app.layout = html.Div([
        
        
        dcc.Upload(
        id='upload-data',
        children=html.Div([
            'Drag and Drop or ',
            html.A('Select Files')
        ]),
        style={
            'width': '100%',
            'height': '60px',
            'lineHeight': '60px',
            'borderWidth': '1px',
            'borderStyle': 'dashed',
            'borderRadius': '5px',
            'textAlign': 'center',
            'margin': '10px'
        },
        # Allow multiple files to be uploaded
        multiple=True
    ),

        dcc.Upload(
        id='upload-data1',
        children=html.Div([
            'Drag and Drop or ',
            html.A('Select Files for Mapping')
        ]),
        style={
            'width': '100%',
            'height': '60px',
            'lineHeight': '60px',
            'borderWidth': '1px',
            'borderStyle': 'dashed',
            'borderRadius': '5px',
            'textAlign': 'center',
            'margin': '10px'
        },
        # Allow multiple files to be uploaded
        multiple=True
    ),
    html.Div(id='output-data-upload'),
    html.Div(id='output-data-upload1'),
        dcc.Checklist(
    id='model-dropdown',        
    options=[
        {'label': 'Template 1', 'value': 'Template 1'},
        {'label': 'Template 2', 'value': 'Template 2'},
        {'label': 'Template 3', 'value': 'Template 3'},
        {'label': 'Template 4', 'value': 'Template 4'},
        {'label': 'Template 5', 'value': 'Template 5'},
        {'label': 'Template 6', 'value': 'Template 6'}
        
    ],
    value = ['Template 1'],
    labelStyle={'display': 'inline-block'}
),               

        
        dcc.Input(id='username', value='Initial Value', type='text'),
        html.Button(id='submit-button', type='submit', children='Submit'),
        html.Div(id='output_div'),

        dcc.Dropdown(id='filename',
                 options=[
                     {'label': i, 'value': i} for i in files
                 ],
                 multi=True
                 ),
        html.Button(id='submit-button1', type='submit', children='Submit'),
        html.Div(id='output_div1'),
                     ])
        
        




@app.callback(Output('output_div', 'children'),
                  [Input('submit-button', 'n_clicks')],
                  [State('username', 'value')],
                  )
def update_output(clicks, input_value):
    if clicks is not None:
        print(input_value)
        
@app.callback(Output('output_div1', 'children'),
                  [Input('submit-button1', 'n_clicks')],
                  [State('filename', 'value')],
                  )
def update_output1(clicks, input_value):
    if clicks is not None:
        print(clicks, input_value)        
        
        
def parse_contents(contents):
    content_type, content_string = contents.split(',')

    decoded = base64.b64decode(content_string)
    try:
#        if 'csv' in filename:
#            # Assume that the user uploaded a CSV file
#            df = pd.read_csv(
#                io.StringIO(decoded.decode('utf-8')))

            # Assume that the user uploaded an excel file
            df = pd.read_excel(io.BytesIO(decoded),"A1")
            #df1 = pd.read_excel(io.BytesIO(decoded),"Datamap")

    except Exception as e:
        print(e)
        return html.Div([
            'There was an error processing this file.'
        ])
            
    return df
#    return html.Div([
#        html.H5(filename),
#        dash_table.DataTable(
#            data=df.to_dict('records'),
#            columns=[{'name': i, 'id': i} for i in df.columns]
#        ),
#
##        dash_table.DataTable(
##            data=df1.to_dict('records'),
##            columns=[{'name': i, 'id': i} for i in df1.columns]
##
##
##        ),
#
#
#
#
#        html.Hr(),
#
#        # For debugging, display the raw contents provided by the web browser
#        html.Div('Raw Content'),
##        html.Pre(contents[0:200] + '...', style={
##            'whiteSpace': 'pre-wrap',
##            'wordBreak': 'break-all'
##        })
#    ])
        
        
        
# file upload function

def parse_contents_df1(contents):
    
    #print(contents)
    
    content_type, content_string = contents.split(',')

    decoded = base64.b64decode(content_string)
    try:
#        if 'csv' in filename:
#            # Assume that the user uploaded a CSV file
#            edu = pd.read_csv(
#                io.StringIO(decoded.decode('ISO-8859-1')))
        
            # Assume that the user uploaded an excel file
            df1 = pd.read_excel(io.BytesIO(decoded),"Datamap")
            
    except Exception as e:
        print(e)
        return None
    #print("This is EDU")
    #print(edu)
    return ((df1))

@app.callback(Output('output-data-upload', 'children'),
              [Input('upload-data', 'contents'),
              Input('username', 'value'),
              Input('model-dropdown','value')])
def update_output(list_of_contents,inputvalue,templatevalue):
    if('Template 1' in templatevalue):
              
        if list_of_contents is not None:
            children = [parse_contents(c) for c in list_of_contents]
            children2 = [parse_contents_df1(c) for c in list_of_contents]
            
            #print((children[0]))
            #print(children)
            df = pd.DataFrame(children[0])
            df1 = pd.DataFrame(children2[0])
            
            #print(df)
            list_rawdata = list(df.columns.values.tolist())
            
            print(df1.head())
            df1.head()
        
            def is_number(s):
                try:
                    float(s)
                    return True
                except ValueError:
                    pass
         
                try:
                    import unicodedata
                    unicodedata.numeric(s)
                    return True
                except (TypeError, ValueError):
                    pass
                
                return False
                    
            for i,l in enumerate(list_rawdata):
                list_rawdata[i]=str(l)       
            
        
            ### Removet the Coolumns which has unnamed in it
            for l in list_rawdata:
                if("Unnamed" in l):
                    list_rawdata.remove(l)
                    
                    ## Converting the list to a String
          
            #### Creating two lists with Questions and their Options and One for the texts    
            list_initial_a = list(df1.iloc[:,0].tolist())
            list_initial_b = list(df1.iloc[:,1].tolist())
            try:
                list_initial_c = list(df1.iloc[:,2].tolist())
            except:
                print("No C")
            for i in range(len(list_initial_a)):
                list_initial_a[i]=str(list_initial_a[i])
            
        
            list_a = []
            list_b = []
            list_c = []
            list_d = []
        
            for i in list_initial_a:
                list_a.append(str(i.split(":")[0]))
                try:
                    list_d.append(str(i.split(":")[1]))
                except:
                    list_d.append("-")
                
            
        
        
        #list_ab = [incom for incom in list_a if str(incom) != 'nan']
        #list_a = list(df1.iloc[:,0].tolist())
            list_b = list(df1.iloc[:,1].tolist())
            try:
                list_c = list(df1.iloc[:,2].tolist())
            except:
                
                print("No C")
        
        #### Treating List_a for : "
        
            
            
        #list_a_type3 = []
        #for i in list_a:
        #    list_a_type3.append(str(i.split(":")[0]))    
            
        ########################    
        #for i in range(0,len(list_a)):
            #print(i)
        #    if((list_a[i] == "nan")):
        #        list_a[i] = list_a[i].replace('nan',' ')
             
        ##################    
            
            list_a1 = []
            list_b1 = []
            list_c1 = []
            list_d1 = []
        
        
        #pd.isna(len(list_a[1]))
        
        #list_a.replace(r'^\s*$', np.nan, regex=True)
        
        
            for i,a in enumerate(list_a):
            #print(pd.isna(list_a[i]))
            #if(len(list_a[i])>1):
            
                if((pd.notnull(a)==True)):
                    list_a1.append(a)
                    list_b1.append(list_b[i])
                    try:
                    
                        list_c1.append(list_c[i])
                        list_d1.append(list_d[i])
                    except:
                        continue 
        
                else:
                    if(pd.notnull(list_b[i])==True):
                        list_a1.append(list_b[i])
                        list_b1.append(list_c[i])
                    
        #df_list = [list_a1,list_b1]
        #df_list_1 =  pd.DataFrame(df_list)
        #df_list_1 = df_list_1.T
        
        ### Creating a list for Questions,Options and Texts
            list_a1_final=[]
            list_b1_final=[]
            list_c1_final = []
            list_d1_final = []
        
            for i, a in enumerate(list_a1):
                if (("type"  not in str(a).lower()) and ("values"  not in str(a).lower() )):
                
                #if(is_number(list_b1[i])==False):
                #print(a)
                    list_a1_final.append(str(a))
                    list_b1_final.append(str(list_b1[i]))    
                    try:
                    
                        list_c1_final.append(str(list_c1[i]))
                        list_d1_final.append(str(list_d1[i]))
                    except:
                        continue
        
        
            list_a2_final=list_a1_final
            list_b2_final=list_b1_final
            try:
            
                list_c2_final = list_c1_final
                list_d2_final = list_d1_final
            except:
                print("No C")
            
            for i,l in enumerate(list_a1_final):
                if(str(l) == 'nan'):
                    list_a2_final[i] = list_b2_final[i]
                    
                    for i,l in enumerate(list_d1_final):
                        if(str(l) != '-'):
                            list_c2_final[i] = list_d2_final[i]
        
        
            new_list_a2_final = []
            
            for j in list_a2_final:
                new_list_a2_final.append(str(j).replace('[','').replace(']',''))
                
        #list_final =  pd.DataFrame(list_a1_final)
            list_final_part_1 =  pd.DataFrame(new_list_a2_final)
            list_final_part_2 =  pd.DataFrame(list_b1_final)
            list_final_part_3 =  pd.DataFrame(list_c2_final)
        
        
        ## Creating a list of all the Question which are not in the Raw Questions List
            notr = []
            for l in (list_a1_final):
                if((l not in list_rawdata) & (l[0].isdigit()!=True) & (len(l)>1)):
                    notr.append(l)
        
        
        
        ############# To indentify MultiSelect##############
            list_final_part_1_list = []
            list_final_part_1_list = list(list_final_part_1.iloc[:,0])
            list_final_part_2_list = []
        
            try:
            
                list_final_part_2_list = list(list_final_part_3.iloc[:,0])
            except:
                list_final_part_2_list = list(list_final_part_2.iloc[:,0])
        
            list_final_part_2_df = pd.DataFrame(list_final_part_2_list)
                
            check_parent = {}
            for l in list_final_part_1_list:
            
                if(((str(l) not in list_rawdata) & ((is_number(l)) == False) &  (len((l)) != 1))):
                    check_parent[l] = "MultiSelectQues"
        
            try:
            
                del check_parent['Open text response']
                del check_parent['Open numeric response']
            except Exception:
                pass
            
        
        
        
        #### Final Data Map ####
        
            final_datamap =  pd.concat([list_final_part_1,list_final_part_2_df],axis = 1)
            final_datamap_df = final_datamap.replace('nan','')
            final_datamap_df = pd.DataFrame(final_datamap_df)
            final_datamap_df.columns = ["Question","Mapping"]
            
        ####### Muli Select Answer Map ####
            new = []
            count = []
            option = []
            original_new = []
            original_new_2 = []
            for key, value in check_parent.items():
                for l in list_final_part_1_list:
                    if(key == str(l)):
                        start_pos = (list_final_part_1_list.index(key))
                        counter  = 0
                        new = []
                        while(list_final_part_1_list[start_pos+counter] != 'nan'):
                            #print(start_pos+counter)
                            if((list_final_part_1_list[start_pos+counter]).isdigit()==True):
                                
                                new.append(list_final_part_1_list[start_pos+counter])
                                original_new.append(list_final_part_1_list[start_pos+counter])
                                original_new_2.append(list_final_part_2_list[start_pos+counter])
            
                            counter = counter + 1
                        original_new.append('')
                        original_new_2.append('')
            
                        count.append(len(new))
            
                option.append(key)
                        
            
            original_new_df = pd.DataFrame(original_new)
            original_new_2_df = pd.DataFrame(original_new_2)
            
            check_parent_ss = {}
            for l in list_final_part_1_list:
                
                if(((str(l) in list_rawdata) & ((is_number(l)) == False) &(str(l) not in check_parent.keys()) & (str(l) not in original_new) )):
                    check_parent_ss[l] = "SingleSelectQues"
            
            
            list_final_part_1_list.append('nan')
            list_final_part_2_list.append('nan')
            
            new_s = []
            count_s = []
            option_s = []
            original_new_s = []
            original_new_2_s = []
            for key, value in check_parent_ss.items():
                for l in list_final_part_1_list:
                    if(key == str(l)):
                        start_pos = (list_final_part_1_list.index(key))
                        counter  = 0
                        new = []
                        while(list_final_part_1_list[start_pos+counter] != 'nan'):
                            #print(start_pos+counter)
                            #if((list_final_part_1_list[start_pos+counter]).isdigit()==False):
                                
                            new_s.append(list_final_part_1_list[start_pos+counter])
                            original_new_s.append(list_final_part_1_list[start_pos+counter])
                            original_new_2_s.append(list_final_part_2_list[start_pos+counter])
            
                            counter = counter + 1
                        original_new_s.append('')
                        original_new_2_s.append('')
            
                        count.append(len(new_s))
            
                option_s.append(key)
            
            original_new_ss_df = pd.DataFrame(original_new_s)
            original_new_ss_2_df = pd.DataFrame(original_new_2_s)
            
            
            
            
            
            
            
            
            ###### Raw Data Column Treatement ##
            
            
            
            new = []
            count = []
            option = []
            original_new_options = []
            original_new_2_options = []
            for key, value in check_parent.items():
                for l in list_final_part_1_list:
                    if(key == str(l)):
                        start_pos = (list_final_part_1_list.index(key))
                        counter  = 1
                        new = []
                            
                        while(list_final_part_1_list[start_pos+counter] != 'nan'):
                            #print(start_pos+counter)
                            if((list_final_part_1_list[start_pos+counter]).isdigit()==False):
                                
                                new.append(list_final_part_1_list[start_pos+counter])
                                original_new_options.append(list_final_part_1_list[start_pos+counter])
                                original_new_2_options.append(list_final_part_2_list[start_pos+counter])
            
                            counter = counter + 1
                        #original_new.append(' ')
                        #original_new_2.append(' ')
            
                        count.append(len(new))
            
                option.append(key)
            
            parents = []
            for key, value in check_parent.items(): 
                parents.append(key)
            
            
            
            new_multiselect = []
            new_multiselect_mapping = []
            i = 0
            for q,p in enumerate(parents):
                print("----")
                for k in range(i,(count[q]+i)):
                    new_multiselect.append(p+"."+original_new_options[k])
                    new_multiselect_mapping.append(p+": "+original_new_2_options[k])
            
                    i = k + 1
                    
                    
                    
            
            #final_multiselect = []
            #original_multiselect = []
            
            #for q,p in enumerate(parents):
            #    for k in range(0,count[q]):
            #        final_multiselect.append(p+".{}".format(k+1))
            #        original_multiselect.append(p)
                    
            #final_multiselect_1 = []       
            #final_multiselect_1 = list(zip(original_new, final_multiselect))
            
            
            ################# Only if Multiselect Questions Format in the Data map are not the same as in the Raw Data
            #original_multiset_rawoptions = []
            #for i,j in enumerate(original_new):
            #    original_multiset_rawoptions.append(original_multiselect[i] + "." + original_new[i])
                
            ###################################################################################################
            
            original_multiset_rawoptions = original_new_options
            #original_multiset_rawoptions = original_new_2_options
            
            #final_multiselect_df = pd.DataFrame(final_multiselect)
            final_new_multiselect_df = pd.DataFrame(new_multiselect)
            final_new_multiselect_mapping_df = pd.DataFrame(new_multiselect_mapping)
            
            original_multiselect_df = pd.DataFrame(original_multiset_rawoptions)
            
            mapping_df = pd.concat([original_multiselect_df,final_new_multiselect_df,final_new_multiselect_mapping_df],axis =1)
            mapping_df.columns = ['A','B','C']
            
            
            list_rawdata = []
            
            list_rawdata = list(df.columns.values.tolist())
            
            
            
            new_listrawdata = list_rawdata
            
            for i,j in enumerate(new_listrawdata):
                #print(j)
                for m in mapping_df.iloc[:,0]:
                    if (str(j) == str(m)):
                        #print(i)
                        new_listrawdata[i] = mapping_df.loc[mapping_df['A'] == m, 'B'].iloc[0]
            
                    
                    
            new_listrawdata_df = pd.DataFrame(new_listrawdata)  
            print("Final Raw Data")
            inputvalue = str(inputvalue) 
            os.chdir((inputvalue))
            new_listrawdata_df.columns = ['Raw Data Columns']
            book = load_workbook('Dynata_v4.xlsx')
            writer = pd.ExcelWriter('Dynata_v4.xlsx', engine='openpyxl') 
            writer.book = book
    
    ## ExcelWriter for some reason uses writer.sheets to access the sheet.
    ## If you leave it empty it will not know that sheet Main is already there
    ## and will create a new sheet.
    
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            final_datamap_df.to_excel(writer, "FinalQuestionMapping11th_2",index = False,header = False)
            writer.save()
            
            wb = load_workbook('Dynata_v4.xlsx', read_only=False, keep_vba=False)
            ws = wb['A1']
    
    # Overwrite Existing data in sheet with a dataframe.
            rows = dataframe_to_rows(new_listrawdata_df, index=False, header=False)
    
            for c_idx, row in enumerate(rows, 1):
                for r_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Save file
            wb.save('Dynata_v4.xlsx')
    
            
            return(html.Div([
            html.H5("Output"),
            dash_table.DataTable(
                data=new_listrawdata_df.to_dict('records'),
                columns=[{'name': i, 'id': i} for i in new_listrawdata_df.columns])]))

    if('Template 2' in templatevalue):
              
        if list_of_contents is not None:
            children = [parse_contents(c) for c in list_of_contents]
            children2 = [parse_contents_df1(c) for c in list_of_contents]
            
            #print((children[0]))
            #print(children)
            df = pd.DataFrame(children[0])
            df1 = pd.DataFrame(children2[0])
            
            #print(df)
            list_rawdata = list(df.columns.values.tolist())
            
            print(df1.head())
            df1.head()
        
            def is_number(s):
                try:
                    float(s)
                    return True
                except ValueError:
                    pass
         
                try:
                    import unicodedata
                    unicodedata.numeric(s)
                    return True
                except (TypeError, ValueError):
                    pass
                
                return False
                    
            for i,l in enumerate(list_rawdata):
                list_rawdata[i]=str(l)       
            
        
            ### Removet the Coolumns which has unnamed in it
            for l in list_rawdata:
                if("Unnamed" in l):
                    list_rawdata.remove(l)
                    
                    ## Converting the list to a String
          
            #### Creating two lists with Questions and their Options and One for the texts    
            list_initial_a = list(df1.iloc[:,0].tolist())
            list_initial_b = list(df1.iloc[:,1].tolist())
            try:
                list_initial_c = list(df1.iloc[:,2].tolist())
            except:
                print("No C")
            for i in range(len(list_initial_a)):
                list_initial_a[i]=str(list_initial_a[i])
            
        
            list_a = []
            list_b = []
            list_c = []
            list_d = []
        
            for i in list_initial_a:
                list_a.append(str(i.split(":")[0]))
                try:
                    list_d.append(str(i.split(":")[1]))
                except:
                    list_d.append("-")
                
            
        
        
        #list_ab = [incom for incom in list_a if str(incom) != 'nan']
        #list_a = list(df1.iloc[:,0].tolist())
            list_b = list(df1.iloc[:,1].tolist())
            try:
                list_c = list(df1.iloc[:,2].tolist())
            except:
                
                print("No C")
        
        #### Treating List_a for : "
        
            
            
        #list_a_type3 = []
        #for i in list_a:
        #    list_a_type3.append(str(i.split(":")[0]))    
            
        ########################    
        #for i in range(0,len(list_a)):
            #print(i)
        #    if((list_a[i] == "nan")):
        #        list_a[i] = list_a[i].replace('nan',' ')
             
        ##################    
            
            list_a1 = []
            list_b1 = []
            list_c1 = []
            list_d1 = []
        
        
        #pd.isna(len(list_a[1]))
        
        #list_a.replace(r'^\s*$', np.nan, regex=True)
        
        
            for i,a in enumerate(list_a):
            #print(pd.isna(list_a[i]))
            #if(len(list_a[i])>1):
            
                if((pd.notnull(a)==True)):
                    list_a1.append(a)
                    list_b1.append(list_b[i])
                    try:
                    
                        list_c1.append(list_c[i])
                        list_d1.append(list_d[i])
                    except:
                        continue 
        
                else:
                    if(pd.notnull(list_b[i])==True):
                        list_a1.append(list_b[i])
                        list_b1.append(list_c[i])
                    
        #df_list = [list_a1,list_b1]
        #df_list_1 =  pd.DataFrame(df_list)
        #df_list_1 = df_list_1.T
        
        ### Creating a list for Questions,Options and Texts
            list_a1_final=[]
            list_b1_final=[]
            list_c1_final = []
            list_d1_final = []
        
            for i, a in enumerate(list_a1):
                if (("type"  not in str(a).lower()) and ("values"  not in str(a).lower() )):
                
                #if(is_number(list_b1[i])==False):
                #print(a)
                    list_a1_final.append(str(a))
                    list_b1_final.append(str(list_b1[i]))    
                    try:
                    
                        list_c1_final.append(str(list_c1[i]))
                        list_d1_final.append(str(list_d1[i]))
                    except:
                        continue
        
        
            list_a2_final=list_a1_final
            list_b2_final=list_b1_final
            try:
            
                list_c2_final = list_c1_final
                list_d2_final = list_d1_final
            except:
                print("No C")
            
            for i,l in enumerate(list_a1_final):
                if(str(l) == 'nan'):
                    list_a2_final[i] = list_b2_final[i]
                    
                    for i,l in enumerate(list_d1_final):
                        if(str(l) != '-'):
                            list_c2_final[i] = list_d2_final[i]
        
        
            new_list_a2_final = []
            
            for j in list_a2_final:
                new_list_a2_final.append(str(j).replace('[','').replace(']',''))
                
        #list_final =  pd.DataFrame(list_a1_final)
            list_final_part_1 =  pd.DataFrame(new_list_a2_final)
            list_final_part_2 =  pd.DataFrame(list_b1_final)
            list_final_part_3 =  pd.DataFrame(list_c2_final)
        
        
        ## Creating a list of all the Question which are not in the Raw Questions List
            notr = []
            for l in (list_a1_final):
                if((l not in list_rawdata) & (l[0].isdigit()!=True) & (len(l)>1)):
                    notr.append(l)
        
        
        
        ############# To indentify MultiSelect##############
            list_final_part_1_list = []
            list_final_part_1_list = list(list_final_part_1.iloc[:,0])
            list_final_part_2_list = []
        
            try:
            
                list_final_part_2_list = list(list_final_part_3.iloc[:,0])
            except:
                list_final_part_2_list = list(list_final_part_2.iloc[:,0])
        
            
            check_parent = {}
            for l in list_final_part_1_list:
            
                if(((str(l) not in list_rawdata) & ((is_number(l)) == False) &  (len((l)) != 1))):
                    check_parent[l] = "MultiSelectQues"
        
            try:
            
                del check_parent['Open text response']
                del check_parent['Open numeric response']
            except Exception:
                pass
            
        
        
        
        #### Final Data Map ####
            list_final_part_2_df = pd.DataFrame(list_final_part_2_list)

            final_datamap =  pd.concat([list_final_part_1,list_final_part_2_df],axis = 1)
            final_datamap_df = final_datamap.replace('nan','')
            final_datamap_df = pd.DataFrame(final_datamap_df)
            final_datamap_df.columns = ["Question","Mapping"]
            
                    
            print("Final Raw Data")

            inputvalue = str(inputvalue) 
            os.chdir((inputvalue))
#            new_listrawdata_df.columns = ['Raw Data Columns']
 


            book = load_workbook('Dynata v1.xlsx')
            writer = pd.ExcelWriter('Dynata v1.xlsx', engine='openpyxl') 
            writer.book = book
    
    ## ExcelWriter for some reason uses writer.sheets to access the sheet.
    ## If you leave it empty it will not know that sheet Main is already there
    ## and will create a new sheet.
    
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            final_datamap_df.to_excel(writer, "FinalQuestionMapping11th_2",index = False,header = False)
            writer.save()
            
    
            
            return(html.Div([
            html.H5("Output")]))


    if('Template 3' in templatevalue):
              
        if list_of_contents is not None:
            children = [parse_contents(c) for c in list_of_contents]
            children2 = [parse_contents_df1(c) for c in list_of_contents]
            
            #print((children[0]))
            #print(children)
            df = pd.DataFrame(children[0])
            df1 = pd.DataFrame(children2[0])
            
            #print(df)
            list_rawdata = list(df.columns.values.tolist())
            
            print(df1.head())
            df1.head()
        
            def is_number(s):
                try:
                    float(s)
                    return True
                except ValueError:
                    pass
         
                try:
                    import unicodedata
                    unicodedata.numeric(s)
                    return True
                except (TypeError, ValueError):
                    pass
                
                return False
                    
            for i,l in enumerate(list_rawdata):
                list_rawdata[i]=str(l)       
            
        
            ### Removet the Coolumns which has unnamed in it
            for l in list_rawdata:
                if("Unnamed" in l):
                    list_rawdata.remove(l)
                    
                    ## Converting the list to a String
          
            #### Creating two lists with Questions and their Options and One for the texts    
            list_initial_a = list(df1.iloc[:,0].tolist())
            list_initial_b = list(df1.iloc[:,1].tolist())
            try:
                list_initial_c = list(df1.iloc[:,2].tolist())
            except:
                print("No C")
            for i in range(len(list_initial_a)):
                list_initial_a[i]=str(list_initial_a[i])
            
        
            list_a = []
            list_b = []
            list_c = []
            list_d = []
        
            for i in list_initial_a:
                list_a.append(str(i.split(":")[0]))
                try:
                    list_d.append(str(i.split(":")[1]))
                except:
                    list_d.append("-")
                
            
        
        
        #list_ab = [incom for incom in list_a if str(incom) != 'nan']
        #list_a = list(df1.iloc[:,0].tolist())
            list_b = list(df1.iloc[:,1].tolist())
            try:
                list_c = list(df1.iloc[:,2].tolist())
            except:
                
                print("No C")
        
        #### Treating List_a for : "
        
            
            
        #list_a_type3 = []
        #for i in list_a:
        #    list_a_type3.append(str(i.split(":")[0]))    
            
        ########################    
        #for i in range(0,len(list_a)):
            #print(i)
        #    if((list_a[i] == "nan")):
        #        list_a[i] = list_a[i].replace('nan',' ')
             
        ##################    
            
            list_a1 = []
            list_b1 = []
            list_c1 = []
            list_d1 = []
        
        
        #pd.isna(len(list_a[1]))
        
        #list_a.replace(r'^\s*$', np.nan, regex=True)
        
        
            for i,a in enumerate(list_a):
            #print(pd.isna(list_a[i]))
            #if(len(list_a[i])>1):
            
                if((pd.notnull(a)==True)):
                    list_a1.append(a)
                    list_b1.append(list_b[i])
                    try:
                    
                        list_c1.append(list_c[i])
                        list_d1.append(list_d[i])
                    except:
                        continue 
        
                else:
                    if(pd.notnull(list_b[i])==True):
                        list_a1.append(list_b[i])
                        list_b1.append(list_c[i])
                    
        #df_list = [list_a1,list_b1]
        #df_list_1 =  pd.DataFrame(df_list)
        #df_list_1 = df_list_1.T
        
        ### Creating a list for Questions,Options and Texts
            list_a1_final=[]
            list_b1_final=[]
            list_c1_final = []
            list_d1_final = []
        
            for i, a in enumerate(list_a1):
                if (("type"  not in str(a).lower()) and ("values"  not in str(a).lower() )):
                
                #if(is_number(list_b1[i])==False):
                #print(a)
                    list_a1_final.append(str(a))
                    list_b1_final.append(str(list_b1[i]))    
                    try:
                    
                        list_c1_final.append(str(list_c1[i]))
                        list_d1_final.append(str(list_d1[i]))
                    except:
                        continue
        
        
            list_a2_final=list_a1_final
            list_b2_final=list_b1_final
            try:
            
                list_c2_final = list_c1_final
                list_d2_final = list_d1_final
            except:
                print("No C")
            
            for i,l in enumerate(list_a1_final):
                if(str(l) == 'nan'):
                    list_a2_final[i] = list_b2_final[i]
                    
                    for i,l in enumerate(list_d1_final):
                        if(str(l) != '-'):
                            list_c2_final[i] = list_d2_final[i]
        
        
            new_list_a2_final = []
            
            for j in list_a2_final:
                new_list_a2_final.append(str(j).replace('[','').replace(']',''))
                
        #list_final =  pd.DataFrame(list_a1_final)
            list_final_part_1 =  pd.DataFrame(new_list_a2_final)
            list_final_part_2 =  pd.DataFrame(list_b1_final)
            list_final_part_3 =  pd.DataFrame(list_c2_final)
        
        
        ## Creating a list of all the Question which are not in the Raw Questions List
            notr = []
            for l in (list_a1_final):
                if((l not in list_rawdata) & (l[0].isdigit()!=True) & (len(l)>1)):
                    notr.append(l)
        
        
        
        ############# To indentify MultiSelect##############
            list_final_part_1_list = []
            list_final_part_1_list = list(list_final_part_1.iloc[:,0])
            list_final_part_2_list = []
        
            try:
            
                list_final_part_2_list = list(list_final_part_3.iloc[:,0])
            except:
                list_final_part_2_list = list(list_final_part_2.iloc[:,0])
        
            
            check_parent = {}
            for l in list_final_part_1_list:
            
                if(((str(l) not in list_rawdata) & ((is_number(l)) == False) &  (len((l)) != 1))):
                    check_parent[l] = "MultiSelectQues"
        
            try:
            
                del check_parent['Open text response']
                del check_parent['Open numeric response']
            except Exception:
                pass
            
        
        
        
        #### Final Data Map ####
            list_final_part_2_df = pd.DataFrame(list_final_part_2_list)

            final_datamap =  pd.concat([list_final_part_1,list_final_part_2_df],axis = 1)
            final_datamap_df = final_datamap.replace('nan','')
            final_datamap_df = pd.DataFrame(final_datamap_df)
            final_datamap_df.columns = ["Question","Mapping"]
            
                    
            print("Final Raw Data")

            inputvalue = str(inputvalue) 
            os.chdir((inputvalue))
#            new_listrawdata_df.columns = ['Raw Data Columns']
 


            book = load_workbook('Dynata v2.xlsx')
            writer = pd.ExcelWriter('Dynata v2.xlsx', engine='openpyxl') 
            writer.book = book
    
    ## ExcelWriter for some reason uses writer.sheets to access the sheet.
    ## If you leave it empty it will not know that sheet Main is already there
    ## and will create a new sheet.
    
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            final_datamap_df.to_excel(writer, "FinalQuestionMapping11th_2",index = False,header = False)
            writer.save()
            
    
            
            return(html.Div([
            html.H5("Output")]))


    if('Template 4' in templatevalue):
              
        if list_of_contents is not None:
            children = [parse_contents(c) for c in list_of_contents]
            children2 = [parse_contents_df1(c) for c in list_of_contents]
            
            #print((children[0]))
            #print(children)
            df = pd.DataFrame(children[0])
            df1 = pd.DataFrame(children2[0])
            
            #print(df)
            list_rawdata = list(df.columns.values.tolist())
            
            print(df1.head())
            df1.head()
        
            def is_number(s):
                try:
                    float(s)
                    return True
                except ValueError:
                    pass
         
                try:
                    import unicodedata
                    unicodedata.numeric(s)
                    return True
                except (TypeError, ValueError):
                    pass
                
                return False
                    
            for i,l in enumerate(list_rawdata):
                list_rawdata[i]=str(l)       
            
        
            ### Removet the Coolumns which has unnamed in it
            for l in list_rawdata:
                if("Unnamed" in l):
                    list_rawdata.remove(l)
                    
                    ## Converting the list to a String
          
            #### Creating two lists with Questions and their Options and One for the texts    
            list_initial_a = list(df1.iloc[:,0].tolist())
            list_initial_b = list(df1.iloc[:,1].tolist())
            try:
                list_initial_c = list(df1.iloc[:,2].tolist())
            except:
                print("No C")
            for i in range(len(list_initial_a)):
                list_initial_a[i]=str(list_initial_a[i])
            
        
            list_a = []
            list_b = []
            list_c = []
            list_d = []
        
            for i in list_initial_a:
                list_a.append(str(i.split(":")[0]))
                try:
                    list_d.append(str(i.split(":")[1]))
                except:
                    list_d.append("-")
                
            
        
        
        #list_ab = [incom for incom in list_a if str(incom) != 'nan']
        #list_a = list(df1.iloc[:,0].tolist())
            list_b = list(df1.iloc[:,1].tolist())
            try:
                list_c = list(df1.iloc[:,2].tolist())
            except:
                
                print("No C")
        
        #### Treating List_a for : "
        
            
            
        #list_a_type3 = []
        #for i in list_a:
        #    list_a_type3.append(str(i.split(":")[0]))    
            
        ########################    
        #for i in range(0,len(list_a)):
            #print(i)
        #    if((list_a[i] == "nan")):
        #        list_a[i] = list_a[i].replace('nan',' ')
             
        ##################    
            
            list_a1 = []
            list_b1 = []
            list_c1 = []
            list_d1 = []
        
        
        #pd.isna(len(list_a[1]))
        
        #list_a.replace(r'^\s*$', np.nan, regex=True)
        
        
            for i,a in enumerate(list_a):
            #print(pd.isna(list_a[i]))
            #if(len(list_a[i])>1):
            
                if((pd.notnull(a)==True)):
                    list_a1.append(a)
                    list_b1.append(list_b[i])
                    try:
                    
                        list_c1.append(list_c[i])
                        list_d1.append(list_d[i])
                    except:
                        continue 
        
                else:
                    if(pd.notnull(list_b[i])==True):
                        list_a1.append(list_b[i])
                        list_b1.append(list_c[i])
                    
        #df_list = [list_a1,list_b1]
        #df_list_1 =  pd.DataFrame(df_list)
        #df_list_1 = df_list_1.T
        
        ### Creating a list for Questions,Options and Texts
            list_a1_final=[]
            list_b1_final=[]
            list_c1_final = []
            list_d1_final = []
        
            for i, a in enumerate(list_a1):
                if (("type"  not in str(a).lower()) and ("values"  not in str(a).lower() )):
                
                #if(is_number(list_b1[i])==False):
                #print(a)
                    list_a1_final.append(str(a))
                    list_b1_final.append(str(list_b1[i]))    
                    try:
                    
                        list_c1_final.append(str(list_c1[i]))
                        list_d1_final.append(str(list_d1[i]))
                    except:
                        continue
        
        
            list_a2_final=list_a1_final
            list_b2_final=list_b1_final
            try:
            
                list_c2_final = list_c1_final
                list_d2_final = list_d1_final
            except:
                print("No C")
            
            for i,l in enumerate(list_a1_final):
                if(str(l) == 'nan'):
                    list_a2_final[i] = list_b2_final[i]
                    
                    for i,l in enumerate(list_d1_final):
                        if(str(l) != '-'):
                            list_c2_final[i] = list_d2_final[i]
        
        
            new_list_a2_final = []
            
            for j in list_a2_final:
                new_list_a2_final.append(str(j).replace('[','').replace(']',''))
                
        #list_final =  pd.DataFrame(list_a1_final)
            list_final_part_1 =  pd.DataFrame(new_list_a2_final)
            list_final_part_2 =  pd.DataFrame(list_b1_final)
            list_final_part_3 =  pd.DataFrame(list_c2_final)
        
        
        ## Creating a list of all the Question which are not in the Raw Questions List
            notr = []
            for l in (list_a1_final):
                if((l not in list_rawdata) & (l[0].isdigit()!=True) & (len(l)>1)):
                    notr.append(l)
        
        
        
        ############# To indentify MultiSelect##############
            list_final_part_1_list = []
            list_final_part_1_list = list(list_final_part_1.iloc[:,0])
            list_final_part_2_list = []
        
            try:
            
                list_final_part_2_list = list(list_final_part_3.iloc[:,0])
            except:
                list_final_part_2_list = list(list_final_part_2.iloc[:,0])
        
            
            check_parent = {}
            for l in list_final_part_1_list:
            
                if(((str(l) not in list_rawdata) & ((is_number(l)) == False) &  (len((l)) != 1))):
                    check_parent[l] = "MultiSelectQues"
        
            try:
            
                del check_parent['Open text response']
                del check_parent['Open numeric response']
            except Exception:
                pass
            
        
        
        
        #### Final Data Map ####
            list_final_part_2_df = pd.DataFrame(list_final_part_2_list)
        
            final_datamap =  pd.concat([list_final_part_1,list_final_part_2_df],axis = 1)
            final_datamap_df = final_datamap.replace('nan','')
            final_datamap_df = pd.DataFrame(final_datamap_df)
            final_datamap_df.columns = ["Question","Mapping"]
            
        ####### Muli Select Answer Map ####
            new = []
            count = []
            option = []
            original_new = []
            original_new_2 = []
            for key, value in check_parent.items():
                for l in list_final_part_1_list:
                    if(key == str(l)):
                        start_pos = (list_final_part_1_list.index(key))
                        counter  = 0
                        new = []
                        while(list_final_part_1_list[start_pos+counter] != 'nan'):
                            #print(start_pos+counter)
                            if((list_final_part_1_list[start_pos+counter]).isdigit()==True):
                                
                                new.append(list_final_part_1_list[start_pos+counter])
                                original_new.append(list_final_part_1_list[start_pos+counter])
                                original_new_2.append(list_final_part_2_list[start_pos+counter])
            
                            counter = counter + 1
                        original_new.append('')
                        original_new_2.append('')
            
                        count.append(len(new))
            
                option.append(key)
                        
            
            original_new_df = pd.DataFrame(original_new)
            original_new_2_df = pd.DataFrame(original_new_2)
            
            check_parent_ss = {}
            for l in list_final_part_1_list:
                
                if(((str(l) in list_rawdata) & ((is_number(l)) == False) &(str(l) not in check_parent.keys()) & (str(l) not in original_new) )):
                    check_parent_ss[l] = "SingleSelectQues"
            
            
            list_final_part_1_list.append('nan')
            list_final_part_2_list.append('nan')
            
            new_s = []
            count_s = []
            option_s = []
            original_new_s = []
            original_new_2_s = []
            for key, value in check_parent_ss.items():
                for l in list_final_part_1_list:
                    if(key == str(l)):
                        start_pos = (list_final_part_1_list.index(key))
                        counter  = 0
                        new = []
                        while(list_final_part_1_list[start_pos+counter] != 'nan'):
                            #print(start_pos+counter)
                            #if((list_final_part_1_list[start_pos+counter]).isdigit()==False):
                                
                            new_s.append(list_final_part_1_list[start_pos+counter])
                            original_new_s.append(list_final_part_1_list[start_pos+counter])
                            original_new_2_s.append(list_final_part_2_list[start_pos+counter])
            
                            counter = counter + 1
                        original_new_s.append('')
                        original_new_2_s.append('')
            
                        count.append(len(new_s))
            
                option_s.append(key)
            
            original_new_ss_df = pd.DataFrame(original_new_s)
            original_new_ss_2_df = pd.DataFrame(original_new_2_s)
            
            
            
            
            
            
            
            
            ###### Raw Data Column Treatement ##
            
            
            
            new = []
            count = []
            option = []
            original_new_options = []
            original_new_2_options = []
            for key, value in check_parent.items():
                for l in list_final_part_1_list:
                    if(key == str(l)):
                        start_pos = (list_final_part_1_list.index(key))
                        counter  = 1
                        new = []
                            
                        while(list_final_part_1_list[start_pos+counter] != 'nan'):
                            #print(start_pos+counter)
                            if((list_final_part_1_list[start_pos+counter]).isdigit()==False):
                                
                                new.append(list_final_part_1_list[start_pos+counter])
                                original_new_options.append(list_final_part_1_list[start_pos+counter])
                                original_new_2_options.append(list_final_part_2_list[start_pos+counter])
            
                            counter = counter + 1
                        #original_new.append(' ')
                        #original_new_2.append(' ')
            
                        count.append(len(new))
            
                option.append(key)
            
            parents = []
            for key, value in check_parent.items(): 
                parents.append(key)
            
            
            
            new_multiselect = []
            new_multiselect_mapping = []
            i = 0
            for q,p in enumerate(parents):
                print("----")
                for k in range(i,(count[q]+i)):
                    new_multiselect.append(p+"."+original_new_options[k])
                    new_multiselect_mapping.append(p+": "+original_new_2_options[k])
            
                    i = k + 1
                    
                    
                    
            
            #final_multiselect = []
            #original_multiselect = []
            
            #for q,p in enumerate(parents):
            #    for k in range(0,count[q]):
            #        final_multiselect.append(p+".{}".format(k+1))
            #        original_multiselect.append(p)
                    
            #final_multiselect_1 = []       
            #final_multiselect_1 = list(zip(original_new, final_multiselect))
            
            
            ################# Only if Multiselect Questions Format in the Data map are not the same as in the Raw Data
            #original_multiset_rawoptions = []
            #for i,j in enumerate(original_new):
            #    original_multiset_rawoptions.append(original_multiselect[i] + "." + original_new[i])
                
            ###################################################################################################
            
            original_multiset_rawoptions = original_new_options
            #original_multiset_rawoptions = original_new_2_options
            
            #final_multiselect_df = pd.DataFrame(final_multiselect)
            final_new_multiselect_df = pd.DataFrame(new_multiselect)
            final_new_multiselect_mapping_df = pd.DataFrame(new_multiselect_mapping)
            
            original_multiselect_df = pd.DataFrame(original_multiset_rawoptions)
            
            mapping_df = pd.concat([original_multiselect_df,final_new_multiselect_df,final_new_multiselect_mapping_df],axis =1)
            mapping_df.columns = ['A','B','C']
            
            
            list_rawdata = []
            
            list_rawdata = list(df.columns.values.tolist())
            
            
            
            new_listrawdata = list_rawdata
            
            for i,j in enumerate(new_listrawdata):
                #print(j)
                for m in mapping_df.iloc[:,0]:
                    if (str(j) == str(m)):
                        #print(i)
                        new_listrawdata[i] = mapping_df.loc[mapping_df['A'] == m, 'B'].iloc[0]
            
                    
                    
            new_listrawdata_df = pd.DataFrame(new_listrawdata)  
            print("Final Raw Data")
            inputvalue = str(inputvalue) 
            os.chdir((inputvalue))
            new_listrawdata_df.columns = ['Raw Data Columns']
            book = load_workbook('Dynata v3.xlsx')
            writer = pd.ExcelWriter('Dynata v3.xlsx', engine='openpyxl') 
            writer.book = book
    
    ## ExcelWriter for some reason uses writer.sheets to access the sheet.
    ## If you leave it empty it will not know that sheet Main is already there
    ## and will create a new sheet.
    
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            final_datamap_df.to_excel(writer, "FinalQuestionMapping12th_2",index = False,header = False)
            writer.save()
            
            wb = load_workbook('Dynata v3.xlsx', read_only=False, keep_vba=False)
            ws = wb['A1']
    
    # Overwrite Existing data in sheet with a dataframe.
            rows = dataframe_to_rows(new_listrawdata_df, index=False, header=False)
    
            for c_idx, row in enumerate(rows, 1):
                for r_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Save file
            wb.save('Dynata v3.xlsx')
    
            
            return(html.Div([
            html.H5("Output"),
            dash_table.DataTable(
                data=new_listrawdata_df.to_dict('records'),
                columns=[{'name': i, 'id': i} for i in new_listrawdata_df.columns])]))



    if('Template 5' in templatevalue):
              
        if list_of_contents is not None:
            children = [parse_contents(c) for c in list_of_contents]
            children2 = [parse_contents_df1(c) for c in list_of_contents]
            
            #print((children[0]))
            #print(children)
            df = pd.DataFrame(children[0])
            df1 = pd.DataFrame(children2[0])
            
            #print(df)
            list_rawdata = list(df.columns.values.tolist())
            
            print(df1.head())
            df1.head()
        
            def is_number(s):
                try:
                    float(s)
                    return True
                except ValueError:
                    pass
         
                try:
                    import unicodedata
                    unicodedata.numeric(s)
                    return True
                except (TypeError, ValueError):
                    pass
                
                return False
                    
            for i,l in enumerate(list_rawdata):
                list_rawdata[i]=str(l)       
            
        
            ### Removet the Coolumns which has unnamed in it
            for l in list_rawdata:
                if("Unnamed" in l):
                    list_rawdata.remove(l)
                    
                    ## Converting the list to a String
          
            #### Creating two lists with Questions and their Options and One for the texts    
            list_initial_a = list(df1.iloc[:,0].tolist())
            list_initial_b = list(df1.iloc[:,1].tolist())
            try:
                list_initial_c = list(df1.iloc[:,2].tolist())
            except:
                print("No C")
            for i in range(len(list_initial_a)):
                list_initial_a[i]=str(list_initial_a[i])
            
        
            list_a = []
            list_b = []
            list_c = []
            list_d = []
        
            for i in list_initial_a:
                list_a.append(str(i.split(":")[0]))
                try:
                    list_d.append(str(i.split(":")[1]))
                except:
                    list_d.append("-")
                
            
        
        
        #list_ab = [incom for incom in list_a if str(incom) != 'nan']
        #list_a = list(df1.iloc[:,0].tolist())
            list_b = list(df1.iloc[:,1].tolist())
            try:
                list_c = list(df1.iloc[:,2].tolist())
            except:
                
                print("No C")
        
        #### Treating List_a for : "
        
            
            
        #list_a_type3 = []
        #for i in list_a:
        #    list_a_type3.append(str(i.split(":")[0]))    
            
        ########################    
        #for i in range(0,len(list_a)):
            #print(i)
        #    if((list_a[i] == "nan")):
        #        list_a[i] = list_a[i].replace('nan',' ')
             
        ##################    
            
            list_a1 = []
            list_b1 = []
            list_c1 = []
            list_d1 = []
        
        
        #pd.isna(len(list_a[1]))
        
        #list_a.replace(r'^\s*$', np.nan, regex=True)
        
        
            for i,a in enumerate(list_a):
            #print(pd.isna(list_a[i]))
            #if(len(list_a[i])>1):
            
                if((pd.notnull(a)==True)):
                    list_a1.append(a)
                    list_b1.append(list_b[i])
                    try:
                    
                        list_c1.append(list_c[i])
                        list_d1.append(list_d[i])
                    except:
                        continue 
        
                else:
                    if(pd.notnull(list_b[i])==True):
                        list_a1.append(list_b[i])
                        list_b1.append(list_c[i])
                    
        #df_list = [list_a1,list_b1]
        #df_list_1 =  pd.DataFrame(df_list)
        #df_list_1 = df_list_1.T
        
        ### Creating a list for Questions,Options and Texts
            list_a1_final=[]
            list_b1_final=[]
            list_c1_final = []
            list_d1_final = []
        
            for i, a in enumerate(list_a1):
                if (("type"  not in str(a).lower()) and ("values"  not in str(a).lower() )):
                
                #if(is_number(list_b1[i])==False):
                #print(a)
                    list_a1_final.append(str(a))
                    list_b1_final.append(str(list_b1[i]))    
                    try:
                    
                        list_c1_final.append(str(list_c1[i]))
                        list_d1_final.append(str(list_d1[i]))
                    except:
                        continue
        
        
            list_a2_final=list_a1_final
            list_b2_final=list_b1_final
            try:
            
                list_c2_final = list_c1_final
                list_d2_final = list_d1_final
            except:
                print("No C")
            
            for i,l in enumerate(list_a1_final):
                if(str(l) == 'nan'):
                    list_a2_final[i] = list_b2_final[i]
                    
                    for i,l in enumerate(list_d1_final):
                        if(str(l) != '-'):
                            list_c2_final[i] = list_d2_final[i]
        
        
            new_list_a2_final = []
            
            for j in list_a2_final:
                new_list_a2_final.append(str(j).replace('[','').replace(']',''))
                
        #list_final =  pd.DataFrame(list_a1_final)
            list_final_part_1 =  pd.DataFrame(new_list_a2_final)
            list_final_part_2 =  pd.DataFrame(list_b1_final)
            list_final_part_3 =  pd.DataFrame(list_c2_final)
        
        
        ## Creating a list of all the Question which are not in the Raw Questions List
            notr = []
            for l in (list_a1_final):
                if((l not in list_rawdata) & (l[0].isdigit()!=True) & (len(l)>1)):
                    notr.append(l)
        
        
        
        ############# To indentify MultiSelect##############
            list_final_part_1_list = []
            list_final_part_1_list = list(list_final_part_1.iloc[:,0])
            list_final_part_2_list = []
        
            try:
            
                list_final_part_2_list = list(list_final_part_3.iloc[:,0])
            except:
                list_final_part_2_list = list(list_final_part_2.iloc[:,0])
        
            
            check_parent = {}
            for l in list_final_part_1_list:
            
                if(((str(l) not in list_rawdata) & ((is_number(l)) == False) &  (len((l)) != 1))):
                    check_parent[l] = "MultiSelectQues"
        
            try:
            
                del check_parent['Open text response']
                del check_parent['Open numeric response']
            except Exception:
                pass
            
        
        
        
        #### Final Data Map ####
            list_final_part_2_df = pd.DataFrame(list_final_part_2_list)
        
            final_datamap =  pd.concat([list_final_part_1,list_final_part_2_df],axis = 1)
            final_datamap_df = final_datamap.replace('nan','')
            final_datamap_df = pd.DataFrame(final_datamap_df)
            final_datamap_df.columns = ["Question","Mapping"]
            
        ####### Muli Select Answer Map ####
            new = []
            count = []
            option = []
            original_new = []
            original_new_2 = []
            for key, value in check_parent.items():
                for l in list_final_part_1_list:
                    if(key == str(l)):
                        start_pos = (list_final_part_1_list.index(key))
                        counter  = 0
                        new = []
                        while(list_final_part_1_list[start_pos+counter] != 'nan'):
                            #print(start_pos+counter)
                            if((list_final_part_1_list[start_pos+counter]).isdigit()==True):
                                
                                new.append(list_final_part_1_list[start_pos+counter])
                                original_new.append(list_final_part_1_list[start_pos+counter])
                                original_new_2.append(list_final_part_2_list[start_pos+counter])
            
                            counter = counter + 1
                        original_new.append('')
                        original_new_2.append('')
            
                        count.append(len(new))
            
                option.append(key)
                        
            
            original_new_df = pd.DataFrame(original_new)
            original_new_2_df = pd.DataFrame(original_new_2)
            
            check_parent_ss = {}
            for l in list_final_part_1_list:
                
                if(((str(l) in list_rawdata) & ((is_number(l)) == False) &(str(l) not in check_parent.keys()) & (str(l) not in original_new) )):
                    check_parent_ss[l] = "SingleSelectQues"
            
            
            list_final_part_1_list.append('nan')
            list_final_part_2_list.append('nan')
            
            new_s = []
            count_s = []
            option_s = []
            original_new_s = []
            original_new_2_s = []
            for key, value in check_parent_ss.items():
                for l in list_final_part_1_list:
                    if(key == str(l)):
                        start_pos = (list_final_part_1_list.index(key))
                        counter  = 0
                        new = []
                        while(list_final_part_1_list[start_pos+counter] != 'nan'):
                            #print(start_pos+counter)
                            #if((list_final_part_1_list[start_pos+counter]).isdigit()==False):
                                
                            new_s.append(list_final_part_1_list[start_pos+counter])
                            original_new_s.append(list_final_part_1_list[start_pos+counter])
                            original_new_2_s.append(list_final_part_2_list[start_pos+counter])
            
                            counter = counter + 1
                        original_new_s.append('')
                        original_new_2_s.append('')
            
                        count.append(len(new_s))
            
                option_s.append(key)
            
            original_new_ss_df = pd.DataFrame(original_new_s)
            original_new_ss_2_df = pd.DataFrame(original_new_2_s)
            
            
            
            
            
            
            
            
            ###### Raw Data Column Treatement ##
            
            
            
            new = []
            count = []
            option = []
            original_new_options = []
            original_new_2_options = []
            for key, value in check_parent.items():
                for l in list_final_part_1_list:
                    if(key == str(l)):
                        start_pos = (list_final_part_1_list.index(key))
                        counter  = 1
                        new = []
                            
                        while(list_final_part_1_list[start_pos+counter] != 'nan'):
                            #print(start_pos+counter)
                            if((list_final_part_1_list[start_pos+counter]).isdigit()==False):
                                
                                new.append(list_final_part_1_list[start_pos+counter])
                                original_new_options.append(list_final_part_1_list[start_pos+counter])
                                original_new_2_options.append(list_final_part_2_list[start_pos+counter])
            
                            counter = counter + 1
                        #original_new.append(' ')
                        #original_new_2.append(' ')
            
                        count.append(len(new))
            
                option.append(key)
            
            parents = []
            for key, value in check_parent.items(): 
                parents.append(key)
            
            
            
            new_multiselect = []
            new_multiselect_mapping = []
            i = 0
            for q,p in enumerate(parents):
                print("----")
                for k in range(i,(count[q]+i)):
                    new_multiselect.append(p+"."+original_new_options[k])
                    new_multiselect_mapping.append(p+": "+original_new_2_options[k])
            
                    i = k + 1
                    
                    
                    
            
            #final_multiselect = []
            #original_multiselect = []
            
            #for q,p in enumerate(parents):
            #    for k in range(0,count[q]):
            #        final_multiselect.append(p+".{}".format(k+1))
            #        original_multiselect.append(p)
                    
            #final_multiselect_1 = []       
            #final_multiselect_1 = list(zip(original_new, final_multiselect))
            
            
            ################# Only if Multiselect Questions Format in the Data map are not the same as in the Raw Data
            #original_multiset_rawoptions = []
            #for i,j in enumerate(original_new):
            #    original_multiset_rawoptions.append(original_multiselect[i] + "." + original_new[i])
                
            ###################################################################################################
            
            original_multiset_rawoptions = original_new_options
            #original_multiset_rawoptions = original_new_2_options
            
            #final_multiselect_df = pd.DataFrame(final_multiselect)
            final_new_multiselect_df = pd.DataFrame(new_multiselect)
            final_new_multiselect_mapping_df = pd.DataFrame(new_multiselect_mapping)
            
            original_multiselect_df = pd.DataFrame(original_multiset_rawoptions)
            
            mapping_df = pd.concat([original_multiselect_df,final_new_multiselect_df,final_new_multiselect_mapping_df],axis =1)
            mapping_df.columns = ['A','B','C']
            
            
            list_rawdata = []
            
            list_rawdata = list(df.columns.values.tolist())
            
            
            
            new_listrawdata = list_rawdata
            
            for i,j in enumerate(new_listrawdata):
                #print(j)
                for m in mapping_df.iloc[:,0]:
                    if (str(j) == str(m)):
                        #print(i)
                        new_listrawdata[i] = mapping_df.loc[mapping_df['A'] == m, 'B'].iloc[0]
            
                    
                    
            new_listrawdata_df = pd.DataFrame(new_listrawdata)  
            print("Final Raw Data")
            inputvalue = str(inputvalue) 
            os.chdir((inputvalue))
            new_listrawdata_df.columns = ['Raw Data Columns']
            book = load_workbook('Dynata_v5.xlsx')
            writer = pd.ExcelWriter('Dynata_v5.xlsx', engine='openpyxl') 
            writer.book = book
    
    ## ExcelWriter for some reason uses writer.sheets to access the sheet.
    ## If you leave it empty it will not know that sheet Main is already there
    ## and will create a new sheet.
    
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            final_datamap_df.to_excel(writer, "FinalQuestionMapping12th_2",index = False,header = False)
            writer.save()
            
            wb = load_workbook('Dynata_v5.xlsx', read_only=False, keep_vba=False)
            ws = wb['A1']
    
    # Overwrite Existing data in sheet with a dataframe.
            rows = dataframe_to_rows(new_listrawdata_df, index=False, header=False)
    
            for c_idx, row in enumerate(rows, 1):
                for r_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Save file
            wb.save('Dynata_v5.xlsx')
    
            
            return(html.Div([
            html.H5("Output"),
            dash_table.DataTable(
                data=new_listrawdata_df.to_dict('records'),
                columns=[{'name': i, 'id': i} for i in new_listrawdata_df.columns])]))



    if('Template 6' in templatevalue):
              
        if list_of_contents is not None:
            children = [parse_contents(c) for c in list_of_contents]
            children2 = [parse_contents_df1(c) for c in list_of_contents]
            
            #print((children[0]))
            #print(children)
            df = pd.DataFrame(children[0])
            df1 = pd.DataFrame(children2[0])
            
            #print(df)
            list_rawdata = list(df.columns.values.tolist())
            
            print(df1.head())
            df1.head()
        
            def is_number(s):
                try:
                    float(s)
                    return True
                except ValueError:
                    pass
         
                try:
                    import unicodedata
                    unicodedata.numeric(s)
                    return True
                except (TypeError, ValueError):
                    pass
                
                return False
                    
            for i,l in enumerate(list_rawdata):
                list_rawdata[i]=str(l)       
            
        
            ### Removet the Coolumns which has unnamed in it
            for l in list_rawdata:
                if("Unnamed" in l):
                    list_rawdata.remove(l)
                    
                    ## Converting the list to a String
          
            #### Creating two lists with Questions and their Options and One for the texts    
            list_initial_a = list(df1.iloc[:,0].tolist())
            list_initial_b = list(df1.iloc[:,1].tolist())
            try:
                list_initial_c = list(df1.iloc[:,2].tolist())
            except:
                print("No C")
            for i in range(len(list_initial_a)):
                list_initial_a[i]=str(list_initial_a[i])
            
        
            list_a = []
            list_b = []
            list_c = []
            list_d = []
        
            for i in list_initial_a:
                list_a.append(str(i.split(":")[0]))
                try:
                    list_d.append(str(i.split(":")[1]))
                except:
                    list_d.append("-")
                
            
        
        
        #list_ab = [incom for incom in list_a if str(incom) != 'nan']
        #list_a = list(df1.iloc[:,0].tolist())
            list_b = list(df1.iloc[:,1].tolist())
            try:
                list_c = list(df1.iloc[:,2].tolist())
            except:
                
                print("No C")
        
        #### Treating List_a for : "
        
            
            
        #list_a_type3 = []
        #for i in list_a:
        #    list_a_type3.append(str(i.split(":")[0]))    
            
        ########################    
        #for i in range(0,len(list_a)):
            #print(i)
        #    if((list_a[i] == "nan")):
        #        list_a[i] = list_a[i].replace('nan',' ')
             
        ##################    
            
            list_a1 = []
            list_b1 = []
            list_c1 = []
            list_d1 = []
        
        
        #pd.isna(len(list_a[1]))
        
        #list_a.replace(r'^\s*$', np.nan, regex=True)
        
        
            for i,a in enumerate(list_a):
            #print(pd.isna(list_a[i]))
            #if(len(list_a[i])>1):
            
                if((pd.notnull(a)==True)):
                    list_a1.append(a)
                    list_b1.append(list_b[i])
                    try:
                    
                        list_c1.append(list_c[i])
                        list_d1.append(list_d[i])
                    except:
                        continue 
        
                else:
                    if(pd.notnull(list_b[i])==True):
                        list_a1.append(list_b[i])
                        list_b1.append(list_c[i])
                    
        #df_list = [list_a1,list_b1]
        #df_list_1 =  pd.DataFrame(df_list)
        #df_list_1 = df_list_1.T
        
        ### Creating a list for Questions,Options and Texts
            list_a1_final=[]
            list_b1_final=[]
            list_c1_final = []
            list_d1_final = []
        
            for i, a in enumerate(list_a1):
                if (("type"  not in str(a).lower()) and ("values"  not in str(a).lower() )):
                
                #if(is_number(list_b1[i])==False):
                #print(a)
                    list_a1_final.append(str(a))
                    list_b1_final.append(str(list_b1[i]))    
                    try:
                    
                        list_c1_final.append(str(list_c1[i]))
                        list_d1_final.append(str(list_d1[i]))
                    except:
                        continue
        
        
            list_a2_final=list_a1_final
            list_b2_final=list_b1_final
            try:
            
                list_c2_final = list_c1_final
                list_d2_final = list_d1_final
            except:
                print("No C")
            
            for i,l in enumerate(list_a1_final):
                if(str(l) == 'nan'):
                    list_a2_final[i] = list_b2_final[i]
                    
                    for i,l in enumerate(list_d1_final):
                        if(str(l) != '-'):
                            list_c2_final[i] = list_d2_final[i]
        
        
            new_list_a2_final = []
            
            for j in list_a2_final:
                new_list_a2_final.append(str(j).replace('[','').replace(']',''))
                
        #list_final =  pd.DataFrame(list_a1_final)
            list_final_part_1 =  pd.DataFrame(new_list_a2_final)
            list_final_part_2 =  pd.DataFrame(list_b1_final)
            list_final_part_3 =  pd.DataFrame(list_c2_final)
        
        
        ## Creating a list of all the Question which are not in the Raw Questions List
            notr = []
            for l in (list_a1_final):
                if((l not in list_rawdata) & (l[0].isdigit()!=True) & (len(l)>1)):
                    notr.append(l)
        
        
        
        ############# To indentify MultiSelect##############
            list_final_part_1_list = []
            list_final_part_1_list = list(list_final_part_1.iloc[:,0])
            list_final_part_2_list = []
        
            try:
            
                list_final_part_2_list = list(list_final_part_3.iloc[:,0])
            except:
                list_final_part_2_list = list(list_final_part_2.iloc[:,0])
        
            
            check_parent = {}
            for l in list_final_part_1_list:
            
                if(((str(l) not in list_rawdata) & ((is_number(l)) == False) &  (len((l)) != 1))):
                    check_parent[l] = "MultiSelectQues"
        
            try:
            
                del check_parent['Open text response']
                del check_parent['Open numeric response']
            except Exception:
                pass
            
        
        
        
        #### Final Data Map ####
            list_final_part_2_df = pd.DataFrame(list_final_part_2_list)
        
            final_datamap =  pd.concat([list_final_part_1,list_final_part_2_df],axis = 1)
            final_datamap_df = final_datamap.replace('nan','')
            final_datamap_df = pd.DataFrame(final_datamap_df)
            final_datamap_df.columns = ["Question","Mapping"]
            
        ####### Muli Select Answer Map ####
            new = []
            count = []
            option = []
            original_new = []
            original_new_2 = []
            for key, value in check_parent.items():
                for l in list_final_part_1_list:
                    if(key == str(l)):
                        start_pos = (list_final_part_1_list.index(key))
                        counter  = 0
                        new = []
                        while(list_final_part_1_list[start_pos+counter] != 'nan'):
                            #print(start_pos+counter)
                            if((list_final_part_1_list[start_pos+counter]).isdigit()==True):
                                
                                new.append(list_final_part_1_list[start_pos+counter])
                                original_new.append(list_final_part_1_list[start_pos+counter])
                                original_new_2.append(list_final_part_2_list[start_pos+counter])
            
                            counter = counter + 1
                        original_new.append('')
                        original_new_2.append('')
            
                        count.append(len(new))
            
                option.append(key)
                        
            
            original_new_df = pd.DataFrame(original_new)
            original_new_2_df = pd.DataFrame(original_new_2)
            
            check_parent_ss = {}
            for l in list_final_part_1_list:
                
                if(((str(l) in list_rawdata) & ((is_number(l)) == False) &(str(l) not in check_parent.keys()) & (str(l) not in original_new) )):
                    check_parent_ss[l] = "SingleSelectQues"
            
            
            list_final_part_1_list.append('nan')
            list_final_part_2_list.append('nan')
            
            new_s = []
            count_s = []
            option_s = []
            original_new_s = []
            original_new_2_s = []
            for key, value in check_parent_ss.items():
                for l in list_final_part_1_list:
                    if(key == str(l)):
                        start_pos = (list_final_part_1_list.index(key))
                        counter  = 0
                        new = []
                        while(list_final_part_1_list[start_pos+counter] != 'nan'):
                            #print(start_pos+counter)
                            #if((list_final_part_1_list[start_pos+counter]).isdigit()==False):
                                
                            new_s.append(list_final_part_1_list[start_pos+counter])
                            original_new_s.append(list_final_part_1_list[start_pos+counter])
                            original_new_2_s.append(list_final_part_2_list[start_pos+counter])
            
                            counter = counter + 1
                        original_new_s.append('')
                        original_new_2_s.append('')
            
                        count.append(len(new_s))
            
                option_s.append(key)
            
            original_new_ss_df = pd.DataFrame(original_new_s)
            original_new_ss_2_df = pd.DataFrame(original_new_2_s)
            
            
            
            
            
            
            
            
            ###### Raw Data Column Treatement ##
            
            
            
            new = []
            count = []
            option = []
            original_new_options = []
            original_new_2_options = []
            for key, value in check_parent.items():
                for l in list_final_part_1_list:
                    if(key == str(l)):
                        start_pos = (list_final_part_1_list.index(key))
                        counter  = 1
                        new = []
                            
                        while(list_final_part_1_list[start_pos+counter] != 'nan'):
                            #print(start_pos+counter)
                            if((list_final_part_1_list[start_pos+counter]).isdigit()==True):
                                
                                new.append(list_final_part_1_list[start_pos+counter])
                                original_new_options.append(list_final_part_1_list[start_pos+counter])
                                original_new_2_options.append(list_final_part_2_list[start_pos+counter])
            
                            counter = counter + 1
                        #original_new.append(' ')
                        #original_new_2.append(' ')
            
                        count.append(len(new))
            
                option.append(key)
            
            parents = []
            for key, value in check_parent.items(): 
                parents.append(key)
            
            
            
            new_multiselect = []
            new_multiselect_mapping = []
            i = 0
            for q,p in enumerate(parents):
                print("----")
                for k in range(i,(count[q]+i)):
                    new_multiselect.append(p+"."+original_new_options[k])
                    new_multiselect_mapping.append(p+": "+original_new_2_options[k])
            
                    i = k + 1
                    
                    
                    
            
            #final_multiselect = []
            #original_multiselect = []
            
            #for q,p in enumerate(parents):
            #    for k in range(0,count[q]):
            #        final_multiselect.append(p+".{}".format(k+1))
            #        original_multiselect.append(p)
                    
            #final_multiselect_1 = []       
            #final_multiselect_1 = list(zip(original_new, final_multiselect))
            
            
            ################# Only if Multiselect Questions Format in the Data map are not the same as in the Raw Data
            #original_multiset_rawoptions = []
            #for i,j in enumerate(original_new):
            #    original_multiset_rawoptions.append(original_multiselect[i] + "." + original_new[i])
                
            ###################################################################################################
            
            original_multiset_rawoptions = original_new_options
            #original_multiset_rawoptions = original_new_2_options
            
            #final_multiselect_df = pd.DataFrame(final_multiselect)
            final_new_multiselect_df = pd.DataFrame(new_multiselect)
            final_new_multiselect_mapping_df = pd.DataFrame(new_multiselect_mapping)
            
            original_multiselect_df = pd.DataFrame(original_multiset_rawoptions)
            
            mapping_df = pd.concat([original_multiselect_df,final_new_multiselect_df,final_new_multiselect_mapping_df],axis =1)
            mapping_df.columns = ['A','B','C']
            
            
            list_rawdata = []
            
            list_rawdata = list(df.columns.values.tolist())
            
            
            
            new_listrawdata = list_rawdata
            
            for i,j in enumerate(new_listrawdata):
                #print(j)
                for m in mapping_df.iloc[:,2]:
                    if (str(j) == str(m)):
                        #print(i)
                        new_listrawdata[i] = mapping_df.loc[mapping_df['C'] == m, 'B'].iloc[0]
            
                    
                    
            new_listrawdata_df = pd.DataFrame(new_listrawdata)  
            print("Final Raw Data")
            inputvalue = str(inputvalue) 
            os.chdir((inputvalue))
            new_listrawdata_df.columns = ['Raw Data Columns']
            
            wb = load_workbook('200409 - BAST - Survey Automation - Pilot v2 Coded Final(Dynata_v6).xlsm', read_only=False, keep_vba=True)
            ws = wb['Answer Key']
            rows_df = dataframe_to_rows(final_datamap_df, index=False, header=True)
            for c_idx, row in enumerate(rows_df, 1):
                for r_idx, value in enumerate(row, 1):
        
                    ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Save file
            wb.save('200409 - BAST - Survey Automation - Pilot v2 Coded Final(Dynata_v6_updated11).xlsm')




            wb = load_workbook('200409 - BAST - Survey Automation - Pilot v2 Coded Final(Dynata_v6).xlsm', read_only=False, keep_vba=True)
            ws = wb['Raw Data']
    
    # Overwrite Existing data in sheet with a dataframe.
            rows = dataframe_to_rows(new_listrawdata_df, index=False, header=False)
            rows_df = dataframe_to_rows(df, index=False, header=True)

            for r_idx, row in enumerate(rows_df, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

    
            for c_idx, row in enumerate(rows, 1):
                for r_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Save file
            wb.save('200409 - BAST - Survey Automation - Pilot v2 Coded Final(Dynata_v6_updated11).xlsm')
    
            
            return(html.Div([
            html.H5("Output"),
            dash_table.DataTable(
                data=new_listrawdata_df.to_dict('records'),
                columns=[{'name': i, 'id': i} for i in new_listrawdata_df.columns])]))






    else:
        
        return(print("Next"))






if __name__ == '__main__':
    app.run_server(debug=True)
